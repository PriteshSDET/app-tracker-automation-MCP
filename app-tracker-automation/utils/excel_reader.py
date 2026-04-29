"""
Excel reader utility for reading test data from Excel files
"""

import openpyxl
from typing import List, Dict, Any
import os
from utils.logger import Logger


class ExcelReader:
    """Excel file reader utility class"""
    
    def __init__(self, file_path: str):
        """Initialize Excel reader with file path"""
        self.file_path = file_path
        self.logger = Logger()
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    def read_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Read data from specific sheet
        
        Args:
            sheet_name: Name of the sheet to read
            
        Returns:
            List of dictionaries containing row data
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            
            sheet = workbook[sheet_name]
            data = []
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Read data rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
                data.append(row_data)
            
            workbook.close()
            self.logger.info(f"Read {len(data)} rows from sheet '{sheet_name}'")
            return data
            
        except Exception as e:
            self.logger.error(f"Failed to read Excel sheet '{sheet_name}': {str(e)}")
            raise
    
    def read_all_sheets(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Read data from all sheets
        
        Returns:
            Dictionary with sheet names as keys and data as values
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            all_data = {}
            
            for sheet_name in workbook.sheetnames:
                all_data[sheet_name] = self.read_sheet(sheet_name)
            
            workbook.close()
            self.logger.info(f"Read data from {len(all_data)} sheets")
            return all_data
            
        except Exception as e:
            self.logger.error(f"Failed to read Excel file: {str(e)}")
            raise
    
    def get_sheet_names(self) -> List[str]:
        """
        Get all sheet names in the workbook
        
        Returns:
            List of sheet names
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
            
        except Exception as e:
            self.logger.error(f"Failed to get sheet names: {str(e)}")
            raise
    
    def get_row_count(self, sheet_name: str) -> int:
        """
        Get number of rows in specific sheet
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Number of data rows (excluding header)
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            
            sheet = workbook[sheet_name]
            row_count = sheet.max_row - 1  # Exclude header row
            workbook.close()
            
            return row_count
            
        except Exception as e:
            self.logger.error(f"Failed to get row count for sheet '{sheet_name}': {str(e)}")
            raise
    
    def find_row_by_value(self, sheet_name: str, column_name: str, search_value: Any) -> Dict[str, Any]:
        """
        Find row by specific column value
        
        Args:
            sheet_name: Name of the sheet
            column_name: Name of the column to search
            search_value: Value to search for
            
        Returns:
            Dictionary containing row data, or empty dict if not found
        """
        try:
            data = self.read_sheet(sheet_name)
            
            for row in data:
                if column_name in row and row[column_name] == search_value:
                    self.logger.info(f"Found row with {column_name}='{search_value}'")
                    return row
            
            self.logger.warning(f"No row found with {column_name}='{search_value}'")
            return {}
            
        except Exception as e:
            self.logger.error(f"Failed to find row by value: {str(e)}")
            raise
    
    def get_column_values(self, sheet_name: str, column_name: str) -> List[Any]:
        """
        Get all values from specific column
        
        Args:
            sheet_name: Name of the sheet
            column_name: Name of the column
            
        Returns:
            List of column values
        """
        try:
            data = self.read_sheet(sheet_name)
            column_values = []
            
            for row in data:
                if column_name in row:
                    column_values.append(row[column_name])
            
            self.logger.info(f"Retrieved {len(column_values)} values from column '{column_name}'")
            return column_values
            
        except Exception as e:
            self.logger.error(f"Failed to get column values: {str(e)}")
            raise


class TestDataLoader:
    """Test data loader for managing Excel test data"""
    
    def __init__(self):
        self.logger = Logger()
        self.readers = {}
    
    def load_users(self, file_path: str = None) -> List[Dict[str, Any]]:
        """Load user test data"""
        from utils.config import Config
        
        file_path = file_path or Config.USERS_FILE
        if file_path not in self.readers:
            self.readers[file_path] = ExcelReader(file_path)
        
        return self.readers[file_path].read_sheet("users")
    
    def load_test_data(self, file_path: str = None) -> Dict[str, List[Dict[str, Any]]]:
        """Load general test data"""
        from utils.config import Config
        
        file_path = file_path or Config.TESTDATA_FILE
        if file_path not in self.readers:
            self.readers[file_path] = ExcelReader(file_path)
        
        return self.readers[file_path].read_all_sheets()
    
    def get_user_by_role(self, role: str) -> Dict[str, Any]:
        """Get user data by role"""
        users = self.load_users()
        
        for user in users:
            if user.get("role") == role:
                return user
        
        raise ValueError(f"No user found with role: {role}")
    
    def get_test_data_by_scenario(self, scenario: str) -> List[Dict[str, Any]]:
        """Get test data by scenario name"""
        test_data = self.load_test_data()
        
        if scenario in test_data:
            return test_data[scenario]
        
        raise ValueError(f"No test data found for scenario: {scenario}")
